#!/usr/bin/env python3
"""Genera docs/crm/Collectionat-CRM.xlsx a partir de docs/crm/leads.csv.

El CSV es la fuente de verdad (versionada en git); el .xlsx es la copia cómoda para
subir a Google Drive y abrir con Google Sheets, ya con filtros, listas desplegables y
resaltado de los leads grado A.

    pip install openpyxl
    python3 scripts/crm-build-xlsx.py
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "docs" / "crm" / "leads.csv"
XLSX_PATH = ROOT / "docs" / "crm" / "Collectionat-CRM.xlsx"

BRAND = "0E7490"  # cyan/petróleo de la marca
ESTADOS = ["nuevo", "verificado", "contactado", "respondió", "demo agendada", "propuesta", "ganado", "perdido"]
GRADOS = ["A", "B", "C"]
MS365 = ["si", "no", "revisar"]
VERTICALES = [
    "Inmobiliarias", "Estudios jurídicos", "Administradoras de consorcios",
    "Escribanías", "Estudios contables", "Distribuidoras / mayoristas",
]
ANCHOS = {
    "empresa": 34, "vertical": 28, "ciudad": 12, "sitio_web": 44, "dominio": 30,
    "proveedor_mail": 18, "cargo_objetivo": 30, "contacto_nombre": 22, "contacto_email": 28,
    "telefono": 18, "estado": 15, "linkedin_busqueda": 30, "linkedin_perfil": 30, "proxima_accion": 32, "ultimo_contacto": 14,
    "fuente": 14, "notas": 70,
}

CRITERIOS = [
    ("Los 4 filtros del ICP", ""),
    ("1. Tamaño", "8 a 150 empleados. Menos de 8 no justifican los USD 4.500; más de 150 ya tienen ERP y comité de compras."),
    ("2. Microsoft 365", "Es el diferencial entero. Una empresa 100% Google Workspace no es lead: es una migración."),
    ("3. Vencimientos que duelen", "Contratos, matrículas, habilitaciones, plazos procesales, expensas, pólizas."),
    ("4. Dueño del desorden", "Alguien cuyo día se arruina cuando falta un papel. Si no existe ese cargo, no hay comprador."),
    ("", ""),
    ("Criterio de fondo", "Collectionat encaja donde el caos es documental y de plazos, no productivo: empresas que administran obligaciones, no que fabrican o mueven cosas."),
    ("", ""),
    ("Verticales A", "Inmobiliarias · Estudios jurídicos · Administradoras de consorcios · Escribanías · Estudios contables · Distribuidoras"),
    ("Verticales B", "Clínicas · Brokers de seguros · Constructoras · Agencias de RRHH · Despachantes de aduana · Colegios privados"),
    ("No perder tiempo", "Retail y gastronomía · Industria y manufactura · Agencias y startups · Freelancers · Corporativos de +300"),
    ("", ""),
    ("Grado A", "Usa Microsoft 365 y entra en tamaño. Acá va el esfuerzo."),
    ("Grado B", "Otro proveedor de mail. Sirve, pero cuesta más."),
    ("Grado C", "Google Workspace, sin dominio propio o fuera de tamaño. Descartar."),
    ("", ""),
    ("Regla de foco", "100 contactos de UN vertical antes de abrir el segundo. Sugerido: estudios jurídicos de La Plata."),
]

FUENTES = [
    ("Estudios jurídicos", "Legal.com.ar (31 estudios en La Plata)", "https://www.legal.com.ar/abogados/estudios-juridicos/la-plata"),
    ("Estudios jurídicos", "AbogadosDe.com.ar", "https://abogadosde.com.ar/buenos-aires/la-plata-147"),
    ("Inmobiliarias", "Zonaprop (181 inmobiliarias en La Plata)", "https://www.zonaprop.com.ar/inmobiliarias-la-plata.html"),
    ("Inmobiliarias", "Inmobúsqueda", "https://www.inmobusqueda.com.ar/inmobiliarias-la-plata.html"),
    ("Inmobiliarias", "Índice La Plata", "https://comercios.indicelaplata.com.ar/inmobiliarias.php"),
    ("Escribanías", "Colegio de Escribanos PBA", "https://www.colescba.org.ar/portal/"),
    ("Escribanías", "Escribanías Argentinas", "https://escribaniasargentinas.com.ar/localidad/la-plata/"),
    ("Administradoras de consorcios", "Registro provincial de administradores", "https://www.gba.gob.ar/dppj/administracion-consorcios"),
]


def sheet_leads(wb: Workbook) -> None:
    with CSV_PATH.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    header, body = rows[0], rows[1:]

    ws = wb.create_sheet("Leads")
    ws.append(header)
    for row in body:
        ws.append(row)

    last_row = len(body) + 1
    for idx, name in enumerate(header, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = ANCHOS.get(name, 16)
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{last_row}"

    for name, opciones in (("estado", ESTADOS), ("grado", GRADOS), ("ms365", MS365), ("vertical", VERTICALES)):
        col = get_column_letter(header.index(name) + 1)
        dv = DataValidation(type="list", formula1=f'"{",".join(opciones)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    grado_col = get_column_letter(header.index("grado") + 1)
    rango = f"{grado_col}2:{grado_col}{max(last_row, 2)}"
    ws.conditional_formatting.add(rango, CellIsRule(
        operator="equal", formula=['"A"'], fill=PatternFill("solid", bgColor="C6F6D5"), font=Font(bold=True)))
    ws.conditional_formatting.add(rango, CellIsRule(
        operator="equal", formula=['"C"'], fill=PatternFill("solid", bgColor="FED7D7")))

    for nombre, etiqueta in (("sitio_web", None), ("linkedin_busqueda", "buscar en LinkedIn")):
        if nombre not in header:
            continue
        col = header.index(nombre) + 1
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=col)
            if not cell.value:
                continue
            url = cell.value
            cell.hyperlink = url
            if etiqueta:
                cell.value = etiqueta
            cell.font = Font(color="0563C1", underline="single")


# Cuánto pesa cada cosa al ordenar la cola de contacto.
PESO_VERTICAL = {
    "Estudios jurídicos": 30, "Estudios contables": 30, "Escribanías": 30,
    "Administradoras de consorcios": 20, "Distribuidoras / mayoristas": 10,
    "Inmobiliarias": 0,
}


def prioridad(fila: dict) -> int:
    """Mayor puntaje = contactar antes."""
    if fila["grado"] == "C" or fila["estado"] in ("descartado", "perdido"):
        return -1
    if not (fila["contacto_email"].strip() or fila["telefono"].strip()):
        return -1  # sin forma de contacto no entra en la cola
    p = 100 if fila["grado"] == "A" else 40
    p += PESO_VERTICAL.get(fila["vertical"], 0)
    if fila["contacto_email"].strip():
        p += 10
    if fila["telefono"].strip():
        p += 5
    if fila["ciudad"].startswith("La Plata"):
        p += 5  # mercado denso y con boca a boca entre colegas
    return p


def sheet_cola(wb: Workbook) -> None:
    with CSV_PATH.open(encoding="utf-8", newline="") as fh:
        filas = list(csv.DictReader(fh))

    cola = sorted(
        ((prioridad(f), f) for f in filas),
        key=lambda par: (-par[0], par[1]["vertical"], par[1]["empresa"]),
    )
    cola = [(p, f) for p, f in cola if p > 0]

    ws = wb.create_sheet("Para contactar", 0)
    encabezado = ["#", "lote", "grado", "empresa", "vertical", "zona", "a quién buscar",
                  "email", "teléfono", "LinkedIn", "por dónde empezar", "contactado"]
    ws.append(encabezado)
    for idx, ancho in enumerate((5, 9, 7, 34, 28, 14, 30, 30, 18, 20, 46, 12), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho
        c = ws.cell(row=1, column=idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 24

    for n, (_, f) in enumerate(cola, start=1):
        lote = f"día {((n - 1) // 15) + 1}"          # 15 por día, como corresponde
        ws.append([n, lote, f["grado"], f["empresa"], f["vertical"], f["ciudad"].split(" / ")[0],
                   f["cargo_objetivo"], f["contacto_email"], f["telefono"], "buscar",
                   f["proxima_accion"], ""])
        fila = ws.max_row
        if f["linkedin_busqueda"]:
            c = ws.cell(row=fila, column=10)
            c.hyperlink = f["linkedin_busqueda"]
            c.font = Font(color="0563C1", underline="single")
        if f["grado"] == "A":
            for col in range(1, len(encabezado) + 1):
                ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="D6F5E0")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezado))}{ws.max_row}"
    dv = DataValidation(type="list", formula1='"sí,no contesta,no interesado"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{ws.max_row + 50}")
    print(f"Cola de contacto: {len(cola)} leads priorizados")


# Invitación para conectar, por rubro. Máximo 300 caracteres: LinkedIn corta.
INVITACIONES = {
    "Estudios jurídicos":
        "Hola [Nombre], veo que llevás la administración del estudio. Trabajo con estudios que "
        "ordenan causas, plazos y documentación en un solo lugar. Me gustaría conectar y conocer "
        "cómo se organizan ustedes. Saludos, Martina.",
    "Escribanías":
        "Hola [Nombre], me interesa mucho cómo trabajan las escribanías el seguimiento de trámites "
        "y documentación. Trabajo en Collectionat, donde ordenamos ese tipo de información. "
        "¿Conectamos? Saludos, Martina.",
    "Administradoras de consorcios":
        "Hola [Nombre], administrar varios consorcios a la vez es un desafío de organización "
        "enorme. Trabajo en Collectionat, ordenando expensas, reclamos y vencimientos por edificio. "
        "Me gustaría conectar. Saludos, Martina.",
    "Estudios contables":
        "Hola [Nombre], los estudios contables manejan más vencimientos por cliente que casi "
        "cualquier otro negocio. Trabajo en Collectionat, ordenando ese calendario. Me gustaría "
        "conectar y conocer cómo lo llevan ustedes. Saludos, Martina.",
    "Distribuidoras / mayoristas":
        "Hola [Nombre], me interesa cómo las distribuidoras siguen las cuentas por cobrar cuando "
        "hay reparto de por medio. Trabajo en Collectionat, ordenando ese tipo de información. "
        "¿Conectamos? Saludos, Martina.",
    "Inmobiliarias":
        "Hola [Nombre], me interesa cómo las inmobiliarias siguen los vencimientos y "
        "actualizaciones de los contratos en administración. Trabajo en Collectionat, ordenando "
        "esa información. ¿Conectamos? Saludos, Martina.",
}


def prioridad_linkedin(fila: dict) -> int:
    """En LinkedIn manda otra lógica: primero los que no se pueden alcanzar por mail."""
    if fila["grado"] == "C" or fila["estado"] in ("descartado", "perdido"):
        return -1
    p = 100 if fila["grado"] == "A" else 40
    if not fila["contacto_email"].strip():
        p += 35  # sin mail, LinkedIn es la única puerta
    p += PESO_VERTICAL.get(fila["vertical"], 0)
    if fila["ciudad"].startswith(("CABA", "GBA", "La Plata")):
        p += 5   # zonas donde hay perfiles cargados de verdad
    return p


def sheet_linkedin(wb: Workbook) -> None:
    with CSV_PATH.open(encoding="utf-8", newline="") as fh:
        filas = list(csv.DictReader(fh))

    cola = sorted(((prioridad_linkedin(f), f) for f in filas),
                  key=lambda par: (-par[0], par[1]["vertical"], par[1]["empresa"]))
    cola = [(p, f) for p, f in cola if p > 0]

    ws = wb.create_sheet("LinkedIn", 1)
    encabezado = ["#", "lote", "grado", "empresa", "vertical", "zona", "a quién buscar",
                  "buscar perfil", "¿tiene mail?", "invitación para copiar",
                  "perfil encontrado", "estado"]
    ws.append(encabezado)
    for idx, ancho in enumerate((5, 9, 7, 34, 28, 14, 30, 16, 12, 78, 34, 16), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho
        c = ws.cell(row=1, column=idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 24

    for n, (_, f) in enumerate(cola, start=1):
        lote = f"día {((n - 1) // 20) + 1}"   # 20 invitaciones por día es el techo sano
        ws.append([n, lote, f["grado"], f["empresa"], f["vertical"], f["ciudad"].split(" / ")[0],
                   f["cargo_objetivo"], "buscar",
                   "sí" if f["contacto_email"].strip() else "NO — solo LinkedIn",
                   INVITACIONES.get(f["vertical"], ""), "", ""])
        fila = ws.max_row
        if f["linkedin_busqueda"]:
            c = ws.cell(row=fila, column=8)
            c.hyperlink = f["linkedin_busqueda"]
            c.font = Font(color="0563C1", underline="single")
        ws.cell(row=fila, column=10).alignment = Alignment(wrap_text=True, vertical="top")
        if f["grado"] == "A":
            for col in range(1, len(encabezado) + 1):
                ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="D6F5E0")
        elif not f["contacto_email"].strip():
            ws.cell(row=fila, column=9).fill = PatternFill("solid", fgColor="FDF0C8")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezado))}{ws.max_row}"
    dv = DataValidation(type="list",
                        formula1='"invitación enviada,aceptada,mensaje enviado,sin respuesta,no aplica"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{ws.max_row + 50}")
    sin_mail = sum(1 for _, f in cola if not f["contacto_email"].strip())
    print(f"Cola de LinkedIn: {len(cola)} leads ({sin_mail} sin mail, arriba de todo)")


def sheet_criterios(wb: Workbook) -> None:
    ws = wb.create_sheet("Criterios ICP")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    for titulo, detalle in CRITERIOS:
        ws.append([titulo, detalle])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")


def sheet_fuentes(wb: Workbook) -> None:
    ws = wb.create_sheet("Fuentes")
    ws.append(["vertical", "directorio", "url"])
    for idx, width in enumerate((30, 44, 70), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND)
    for vertical, nombre, url in FUENTES:
        ws.append([vertical, nombre, url])
        cell = ws.cell(row=ws.max_row, column=3)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_leads(wb)
    sheet_cola(wb)
    sheet_linkedin(wb)
    sheet_criterios(wb)
    sheet_fuentes(wb)
    wb.save(XLSX_PATH)
    print(f"Generado {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
